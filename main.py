import VR_Security_with_GPT as tool
import pandas as pd
import openpyxl
import time


# 读取单元格每行内容，根据第一列的文件名找到gpt_app中的apk，根据第二列的游戏描述，生成对应的非法权限作为回答并保存在第三列，读取的xml文件存在第四列
# 单元格第一列内容应为对应apk名，且去掉.apk。注意名称要去掉空格。
# 单元格第二列内容为对应apk的描述。


if __name__ == "__main__":

    file_path = 'app_permission.xlsx'  # 替换为实际的文件路径
    sheet_name = 'GPT_vs_GT'

    # 读
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    selected_data = df.loc[:, 0:1]
    df = None # 关闭读
    # 写初始化
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 打印选定的数据。index是之前读取的时候的index，从0开始。但sheet.cell中索引是从1开始的。
    for index, row in selected_data.iterrows():
        try:
            print("\n\n#################################### "+str(index)+" ############################################################################################################################################################\n\n")
            # 如果已经写过就跳过
            if sheet.cell(row=index+1, column=3).value != "" and sheet.cell(row=index+1, column=4).value != "" and sheet.cell(row=index+1, column=3).value is not None and sheet.cell(row=index+1, column=4).value is not None:
                print("already wrote.")
                continue
            # 提问
            apk_src = "D:\\ict\\My_Project\\Security\\reverse\\reverse_app\\gpt_app\\"+row[0]+".apk"
            description = row[1]
            answer,xml_string = tool.analyzeOneApk(apk_src,description)
            # 保存结果
            sheet.cell(row=index+1, column=3).value = answer
            sheet.cell(row=index+1, column=4).value = xml_string
            workbook.save(file_path)

            for i in range(21):
                time.sleep(1)
                print("sleep:"+str(i+1)+"s")

        except Exception as e:
            print(f"Error occurred: {str(e)}")










